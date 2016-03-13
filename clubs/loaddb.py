__author__ = 'ayoubamnour'


from openpyxl import load_workbook
from clubs.models import Club, Player, Manager, PreviousClubs

def insert_club():

    increment =2
    wb2 = load_workbook('footbal_teams.xlsx')
    sheet = wb2.get_sheet_by_name('Sheet1')

    for row in sheet.iter_rows(row_offset=1):
        club_name = sheet[str('R'+str(increment))].value
        # manager = sheet[str('C'+str(increment))].value
        city = sheet[str('T'+str(increment))].value
        Arena = sheet[str('U'+str(increment))].value

        club = Club(club_name=club_name,city=city,Arena=Arena)
        club.save()

        increment +=1


def insert_players():

    increment =2
    wb2 = load_workbook('footbal_teams.xlsx')
    sheet = wb2.get_sheet_by_name('Sheet1')

    for row in sheet.iter_rows(row_offset=1):
        number = sheet[str('A'+str(increment))].value
        name = sheet[str('B'+str(increment))].value
        # nationality = sheet[str('C'+str(increment))].value
        position = sheet[str('D'+str(increment))].value
        height = sheet[str('F'+str(increment))].value
        # weight = sheet[str('G'+str(increment))].value
        date_of_birth = sheet[str('H'+str(increment))].value
        # previous_club = sheet[str('I'+str(increment))].value
        club = sheet[str('J'+str(increment))].value

        player = Player(club=club,number=number,name=name,position=position,height=height,date_of_birth=date_of_birth)
        player.save()

        increment +=1


def insert_manager():

    increment =2
    wb2 = load_workbook('footbal_teams.xlsx')
    sheet = wb2.get_sheet_by_name('Sheet1')

    for row in sheet.iter_rows(row_offset=1):
        # club_name = sheet[str('X'+str(increment))].value
        club_manager = sheet[str('Y'+str(increment))].value

        manager = Manager(name=club_manager)
        manager.save()

        increment +=1


def insert_previous_clubs():

    increment =2
    wb2 = load_workbook('footbal_teams.xlsx')
    sheet = wb2.get_sheet_by_name('Sheet1')

    for row in sheet.iter_rows(row_offset=1):
        name = sheet[str('B'+str(increment))].value
        previous_club = sheet[str('I'+str(increment))].value

        previousClub = PreviousClubs(player=name,club=previous_club)
        previousClub.save()

        increment +=1




import os
if __name__ == '__main__':
    # os.environ.setdefault("DJANGO_SETTINGS_MODULE", "FootbalRestAPI.settings")
    # insert_club()
    insert_manager()
    # insert_players()
    # insert_previous_clubs()
    # previousClub = PreviousClubs(player="ayoub",club="Real Madrid")
    # previousClub.save()